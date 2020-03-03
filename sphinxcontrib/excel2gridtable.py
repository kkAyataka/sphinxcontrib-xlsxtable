import unicodedata
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


def get_string_width(text: str):
    width = 0
    for t in text:
        if unicodedata.east_asian_width(t) in 'FWA':
            width += 2
        else:
            width += 1
    return width

def get_max_width(lines):
    max_w = 0
    for l in lines:
        max_w = max(max_w, get_string_width(l))
    return max_w

def count_line(values):
    count = len(values)
    if '|' in ''.join(values):
        count += 1
    return count

class TableCell:
    is_merged_top = False
    is_merged_left = False

    def __init__(self, row, column, value):
        values = []
        if value != None:
            values = f'{value}'.split('\n')

        self.row = row
        self.column = column
        self.values = values
        self.line_count = count_line(values)
        self.width = get_max_width(values)

def get_padding(count, max_count):
    padding = ''
    for _ in range(max_count - count):
        padding += ' '
    return padding

def get_rule(colmuns, is_head=False):
    line_str = ''
    for cell in colmuns:
        line_str += '+'
        for _ in range(cell.width + 2):
            if cell.is_merged_top:
                line_str += ' '
            elif is_head:
                line_str += '='
            else:
                line_str += '-'

    line_str += '+'
    return line_str

def gen_reST_grid_table_lines(filename, header_rows, sheetname=None):
    wb = load_workbook(
        filename=filename,
        read_only=False, # Can not get merged cell information if read_only is True
        keep_vba=False,
        data_only=True,
        keep_links=False
        )

    # select target sheet
    try:
        ws = wb[sheetname]
    except:
        ws = wb.active

    # parse cell info
    table_cells = []
    for r in range(1, ws.max_row + 1):
        # appebd array for row
        table_cells.append([])

        # get line count in the row
        for c  in range(1, ws.max_column + 1):
            tc = TableCell(r, c, ws.cell(r, c).value)
            table_cells[r - 1].append(tc)

    # adjust line count
    row_count = len(table_cells)
    col_count = len(table_cells[0])
    for r in range(row_count):
        max_line_count = 0
        for c in range(col_count):
            max_line_count = max(max_line_count, table_cells[r][c].line_count)

        for c in range(col_count):
            table_cells[r][c].line_count = max_line_count

    # adjust width
    for c in range(col_count):
        max_width = 0
        for r in range(row_count):
            max_width = max(max_width, table_cells[r][c].width)

        for r in range(row_count):
            table_cells[r][c].width = max_width

    # adjust merged cell info
    for mrange in ws.merged_cell_ranges:
        left = mrange.bounds[0] - 1
        top = mrange.bounds[1] - 1
        right = mrange.bounds[2] - 1
        bottom = mrange.bounds[3] - 1
        for c in range(left, right + 1):
            for r in range(top, bottom + 1):
                table_cells[r][c].is_merged_top = (r != top)
                table_cells[r][c].is_merged_left = (c != left)

    # gen lines
    grid_table_lines = []
    for r, cols in enumerate(table_cells):
        if r == header_rows and header_rows > 0:
            grid_table_lines.append(get_rule(cols, True))
        else:
            grid_table_lines.append(get_rule(cols))

        line_count = cols[0].line_count
        for l in range(line_count):
            line_str = ''
            for cell in cols:
                line_str += '| ' if cell.is_merged_left == False else '  '
                if len(cell.values) > l:
                    line_str += f'{cell.values[l]}'
                    line_str += get_padding(get_string_width(cell.values[l]), cell.width)
                    line_str += ' '
                else:
                    line_str += get_padding(0, cell.width)
                    line_str += ' '
            line_str += '|'

            grid_table_lines.append(line_str)

    grid_table_lines.append(get_rule(cols))
    return grid_table_lines

def draw_reST_grid_table(filename, header_rows):
    lines = gen_reST_grid_table_lines(filename, header_rows)
    for l in lines:
        print(l)

if __name__ == '__main__':
    filename = 'sample.xlsx'
    header_rows = 1
    draw_reST_grid_table(filename, header_rows)
