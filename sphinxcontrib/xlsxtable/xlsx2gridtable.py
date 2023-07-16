import argparse
import os
import unicodedata
from openpyxl import load_workbook
from .workbook.workbook import Workbook as ImageWorkbook

def parse_index_str(index_str: str) -> int:
    """ Gets index number form Excel column letter (1-based).
        1 from A, 26 from Z and 52 from AZ.
    """
    index_str = index_str.upper()
    if index_str[0] in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
        index_str_len = len(index_str)
        index = 0
        for i in range(index_str_len):
            index += (ord(index_str[i]) - ord('A') + 1) * pow(26, index_str_len - 1 - i)
        return index
    else:
        return int(index_str)

def parse_indexes_str(indexes_str: str) -> list:
    """ Gets indexes numbers from string (1-based).
        1, 2, 4 and 6 from '1-2, D-F'.
    """
    indexes = []
    range_indexes_strs = indexes_str.replace(',', ' ').split()
    for range_str in range_indexes_strs:
        if '-' in range_str:
            range_str_els = range_str.split('-')
            begin = parse_index_str(range_str_els[0])
            end = parse_index_str(range_str_els[1])
            indexes += list(range(begin, end + 1))
        else:
            indexes.append(parse_index_str(range_str))

    return list(set(indexes))

def get_use_indexes(min_index: int, max_index: int, includes_str: str, excludes_str: str):
    """ Gets valid indexes (1-based).
    """
    includes = set(range(min_index, max_index + 1))
    if includes_str is not None and len(includes_str) > 0:
        includes = set(parse_indexes_str(includes_str))

    excludes = set([])
    if excludes_str is not None and len(excludes_str) > 0:
        excludes = set(parse_indexes_str(excludes_str))

    return list(includes - excludes)

def get_string_width(text: str):
    width = 0
    for t in text:
        if unicodedata.east_asian_width(t) in 'FW':
            width += 2
        else:
            width += 1
    return width

def get_max_width(lines):
    max_w = 0
    for l in lines:
        max_w = max(max_w, get_string_width(l))
    return max_w

def count_line(values):
    count = len(values)
    if '|' in ''.join(values):
        count += 1
    return count

class TableCell:
    is_merged_top = False
    is_merged_left = False

    def __init__(self, row, column, value):
        self.row = row
        self.column = column
        self.set_value(value)

    def set_value(self, value: str):
        values = []
        if value != None:
            values = f'{value}'.split('\n')

        self.values = values
        self.line_count = max(count_line(values), 1)
        self.width = get_max_width(values)

def get_cell(table_cells: list[list], row: int, column: int):
    for cols in table_cells:
        for cell in cols:
            if cell.row == row and cell.column == column:
                return cell
    return None

def get_padding(count, max_count):
    padding = ''
    for _ in range(max_count - count):
        padding += ' '
    return padding

def get_rule(colmuns, is_head=False, is_end=False, is_top=False):
    line_str = ''
    for c, cell in enumerate(colmuns):
        if c != 0 and cell.is_merged_left and cell.is_merged_top:
            line_str += ' '
        else:
            line_str += '+'
        for _ in range(cell.width + 2):
            if is_top:
                line_str += '-'
            elif cell.is_merged_top and not is_end:
                line_str += ' '
            elif is_head:
                line_str += '='
            else:
                line_str += '-'

    line_str += '+'
    return line_str

def gen_reST_grid_table_lines(
    file: str,
    fullpath: str,
    header_rows=0,
    sheetname=None,
    start_row=1,
    start_column=1,
    include_rows=None,
    exclude_rows=None,
    include_columns=None,
    exclude_columns=None):
    """
    :param file: The :file: parameter value in the directive
    :param fullpath: The full path for the :file: parameter file
    """

    # Parse workbook
    wb = load_workbook(
        filename=fullpath,
        read_only=False, # Can not get merged cell information if read_only is True
        keep_vba=False,
        data_only=True,
        keep_links=False
        )

    # select target sheet
    try:
        ws = wb[sheetname]
    except:
        ws = wb.active

    # Parse for images
    img_wb = None
    img_ws = None
    try:
        img_wb = ImageWorkbook.parse(fullpath)
        img_ws = img_wb.get_sheet(ws.title)
    except:
        pass
    img_min_from_row = img_ws.drawing.min_from_row + 1 if img_ws != None and img_ws.drawing != None else 1048576
    img_max_from_row = img_ws.drawing.max_from_row + 1 if img_ws != None and img_ws.drawing != None else 1
    img_min_from_col = img_ws.drawing.min_from_col + 1 if img_ws != None and img_ws.drawing != None else 16384
    img_max_from_col = img_ws.drawing.max_from_col + 1 if img_ws != None and img_ws.drawing != None else 1

    # rows / columns
    min_row = min(ws.min_row, img_min_from_row)
    max_row = max(ws.max_row, img_max_from_row)
    min_col = min(ws.min_column, img_min_from_col)
    max_col = max(ws.max_column, img_max_from_col)
    offset_row = max(min_row, start_row)
    offset_col = max(min_col, start_column)

    use_rows = get_use_indexes(min_row, max_row, include_rows, exclude_rows)
    use_cols = get_use_indexes(min_col, max_col, include_columns, exclude_columns)

    # parse cell info
    table_cells = []
    for r in range(offset_row, max_row + 1):
        if r in use_rows:
            # appebd array for row
            table_cells.append([])

            # get line count in the row
            r_index = len(table_cells) - 1
            for c in range(offset_col, max_col + 1):
                if c in use_cols:
                    tc = TableCell(r, c, ws.cell(r, c).value)
                    table_cells[r_index].append(tc)

    # Handles images, embeds image directive
    if img_ws != None and img_ws.drawing != None:
        # Make images directory from xlsx file
        base_dir = os.path.dirname(fullpath)
        dir_name = os.path.basename(fullpath) + '.media'
        img_dir = os.path.join(base_dir, dir_name)
        if os.path.exists(img_dir) == False:
            os.mkdir(img_dir)
        # Write out image files
        for m in img_wb.media:
            file_path = os.path.join(img_dir, os.path.basename(m))
            with open(file_path, 'wb') as fs:
                fs.write(img_wb.media[m])
        # Add image directives in cells
        for anchor in img_ws.drawing.two_cell_anchors:
            r = anchor.from_pt.row - (offset_row - 1)
            c = anchor.from_pt.col - (offset_col - 1)
            table_cells[r][c].set_value(f'.. image:: {file}.media/{os.path.basename(anchor.image_path)}')

    # adjust line count
    row_count = len(table_cells)
    col_count = len(table_cells[0])
    for r in range(row_count):
        max_line_count = 0
        for c in range(col_count):
            max_line_count = max(max_line_count, table_cells[r][c].line_count)

        for c in range(col_count):
            table_cells[r][c].line_count = max_line_count

    # adjust width
    for c in range(col_count):
        max_width = 0
        for r in range(row_count):
            max_width = max(max_width, table_cells[r][c].width)

        for r in range(row_count):
            table_cells[r][c].width = max_width

    # adjust merged cell info
    for mrange in ws.merged_cells.ranges:
        left = mrange.bounds[0]
        top = mrange.bounds[1]
        right = mrange.bounds[2]
        bottom = mrange.bounds[3]
        for c in range(left, right + 1):
            for r in range(top, bottom + 1):
                cell = get_cell(table_cells, r, c)
                if cell is not None:
                    cell.is_merged_top = (r != top)
                    cell.is_merged_left = (c != left)

    # gen lines
    grid_table_lines = []
    for r, cols in enumerate(table_cells):
        if r == 0:
            grid_table_lines.append(get_rule(cols, is_top=True))
        elif r == header_rows:
            grid_table_lines.append(get_rule(cols, is_head=True))
        else:
            grid_table_lines.append(get_rule(cols))

        line_count = cols[0].line_count
        for l in range(line_count):
            line_str = ''
            for c, cell in enumerate(cols):
                line_str += '| ' if c == 0 or cell.is_merged_left == False else '  '
                if len(cell.values) > l:
                    line_str += f'{cell.values[l]}'
                    line_str += get_padding(get_string_width(cell.values[l]), cell.width)
                    line_str += ' '
                else:
                    line_str += get_padding(0, cell.width)
                    line_str += ' '
            line_str += '|'

            grid_table_lines.append(line_str)

    grid_table_lines.append(get_rule(cols, is_end=True))
    return grid_table_lines

def draw_reST_grid_table(
        file, fullpath, header_rows, sheet, start_row, start_column,
        include_rows, exclude_rows,
        include_columns, exclude_columns):
    lines = gen_reST_grid_table_lines(file, fullpath, header_rows, sheet, start_row, start_column,
        include_rows, exclude_rows, include_columns, exclude_columns)

    text ='\n'.join(lines)
    print(text)
    return text

def main():
    p = argparse.ArgumentParser(description='Grid Table String Generator')
    p.add_argument('--header-rows', type=int, default=0, help='Header rows')
    p.add_argument('--sheet', type=str, help='Target sheet name')
    p.add_argument('--start-row', type=int, default=1, help='Start row')
    p.add_argument('--start-column', type=int, default=1, help='Start colmun')
    p.add_argument('--include-rows', type=str, default=None, help='Specify included rows')
    p.add_argument('--exclude-rows', type=str, default=None, help='Specify excluded rows')
    p.add_argument('--include-columns', type=str, default=None, help='Specify included columns')
    p.add_argument('--exclude-columns', type=str, default=None, help='Specify excluded columns')
    p.add_argument('file', type=str, help='Target Excel file path')

    args = p.parse_args()

    draw_reST_grid_table(args.file, args.file, args.header_rows, args.sheet,
        args.start_row, args.start_column,
        args.include_rows, args.exclude_rows,
        args.include_columns, args.exclude_columns)

if __name__ == '__main__':
    main()
